import openpyxl
import xlwings as xw
import json

def HealthClub_Workers():
    path = "C:/Users/roman/OneDrive/Рабочий стол/Health Club/WorkedOut.xlsx"

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    wb=xw.Book('WorkedOut.xlsx')
    i = 3
    reactive = []
    activeBadWorkers = []
    didTheyDonePlan = []
    
    while i <= 6:
        WorkerName = sheet_obj.cell(row = i+1, column = 1).value
        dataofworker = wb.sheets['Health club']
        dataofworkers = dataofworker.range(f'B{i+1}:D{i+1}').options(index = False).value
        total = dataofworker.range(f'L{i+1}').options(index = False).value
        
        WorkerStatusing = {
                f"{WorkerName}":min(dataofworkers)
                    }
        
        activeBadWorkers.append(WorkerStatusing)
        i += 1
        
        if total > 0: absence = "Yes"
        else: absence = "No"
        
        didPlan = {
                f"{WorkerName}":absence
        }
        
        WorkerStatus = {
                f"{WorkerName}":[{"Maximum Wage":max(dataofworkers)},
                        {"Minimum Wage":min(dataofworkers)},
                        {"Total":total},
                        {"Did the plan?":absence}
                    ]}
    
        reactive.append(WorkerStatus)
        didTheyDonePlan.append(didPlan)
        # print(f"{WorkerName} , {absence}")

    rev1 = []
    rev = []
    a = 0
    while a <= 3:
        valueted = ', '.join(str(x) for x in activeBadWorkers[a].values())
        rev1.append(valueted)
        for key,v in activeBadWorkers[a].items():
            rev.append(key)
   
        a += 1
    
    rev12 = []
    rev2 = []
    a2 = 0
    while a2 <= 3:
        valueted = ', '.join(str(x) for x in didTheyDonePlan[a2].values())
        rev12.append(valueted)
        for key,v in didTheyDonePlan[a2].items():
            rev2.append(key)      
   
        a2 += 1
            
    dicts = {}
    
    for dictionaryValues in range(len(rev)):
        dicts[rev[dictionaryValues]] = rev1[dictionaryValues]
        
    diction = {}
    
    for dictionaryValues in range(len(rev2)):
        diction[rev2[dictionaryValues]] = rev12[dictionaryValues]

    # need for aadditionals work
    
    sortedDic = {}
    sortedKeys = sorted(dicts, key=dicts.get)
    for w in sortedKeys:
        sortedDic[w] = dicts[w]
    
    
    free = []
    for t in sortedDic:
        free.append(t)
    
    plandic = []
    abcd = 0
    while abcd <= 1:
        if diction[free[abcd]] == "Yes": plan = "done"
        else: plan = "none"
        abcd += 1
        plandic.append(plan)  
    
    darknesS = []
    h = 0
    while h <= 1:
        improveYourSelf = {
            f"{free[h]}":plandic[h]
        }
        darknesS.append(improveYourSelf)
        
        # print(darknesS[h][f"{free[h]}"])
        if darknesS[h][f"{free[h]}"] == "done":
            # ("So far as, one of the worker didn't reached the plan. ")
            print(f"The worker {free[h]} done his plan work. However, we have couple of advice for him")
        else:  print(f"The worker {free[h]} had a troubles to achive a goal with plan, thats not a problem, lets try amplify him")

        h += 1 
    
    
    # Perhaps here will be a bug, becouse of if statement (up) 
    
    def PrepareForFront():
        PrepareForFront = {
            "React Text 1":f"The workers {free[0]}, {free[1]} had the lowest wage. Which contain {sortedDic[free[0]]} , {sortedDic[free[1]]} accordingly",
            "React Text 2":f"The worker {free[0]} {plandic[0]} plan and worker {free[1]} {plandic[1]} the plan ",
            "React Text 3":f"The worker {free[h]} done his plan work. However, we have couple of advice for him",
            "React Text 4":f"The worker {free[h]} had a troubles to achive a goal with plan, thats not a problem, lets try amplify him",
            "Excellent 1":f"The worker {WorkerName} had lowest wages {min(dataofworkers)} in December, keep doing more!",
            "Excellent 2":f"The worker {WorkerName} had highest wages {max(dataofworkers)} in September, excellent!",
        }
        
        Prepare = json.dumps(PrepareForFront)
        Front = open("PrepareForFront.json", "w")
        Front.write(Prepare)
        Front.close()
    
    PrepareForFront()
    
    # need for aadditionals work
    
    # Transportiration in the new folder
    
    
    # Transportiration in the new folder
    
    jsonString = json.dumps(reactive)
    jsonFile = open("Statistics.json", "w")
    jsonFile.write(jsonString)
    jsonFile.close()
    
def HealthClub_Improvements():
    path = "C:/Users/roman/OneDrive/Рабочий стол/Health Club/Improvements.xlsx"

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    wb=xw.Book('Improvements.xlsx')

    i = 3
    listing = []
    while i <= 6:
        WorkerName = sheet_obj.cell(row = i+1, column = 1).value
        dataofworker = wb.sheets['Improvements']
        study = dataofworker.range('B3').options(index = False).value
        additional = dataofworker.range('C3').options(index = False).value
        skills = dataofworker.range(f'B{i+1}:C{i+1}').options(index = False).value
        i += 1

        if skills.index('Yes') == 0: maintain = study
        if skills.index('Yes') == 1: maintain = additional
        
        skilling = {
            f"{WorkerName}":[{
                "Skills":maintain
                            }]
            }
        
        listing.append(skilling)
        
        # print(f"The worker {WorkerName}, have a ability {study}.")
        # print(f"The worker {WorkerName}, have a ability {additional}.")
        
    jsonString = json.dumps(listing)
    jsonFile = open("Skilling.json", "w")
    jsonFile.write(jsonString)
    jsonFile.close()
    
def HealthClub_Quiz():
    result = None
    perhapsAnswer = None
    WorkerName = "Romash"

    Quiz = {
        "Topics":{
            "Personality":{
                f"Is worker {WorkerName} an active person?":result,
                f"How many special ability worker {WorkerName} have":result,
                f"Is {WorkerName} talkative? ":result,
                f"How many times {WorkerName} have a vacation?":result
            },
            "Carier":{
                "How do you think, can season influent of profit?":result,
                "Would you like to requilifing some workers? ":result,
                "How many amount of time you require to find some captivating course for your mates?":result,
                "Try to input new services on the month?":result
            },
            "Plan Generation":{
                "Whats will be you reaction, if you decline of Plan level?":result,
                "Lets add some services to the worker {WorkerName}. Because new services can apply ... ":result,
                "To supply workers more customers, we need to amplify quilification level. For instance, try to advice {WorkerName} new books or videos. Try to unity a team, gathering they are all together":result,
                "In my mind i got a story, about girl. She doesn't like programming as well, but she adore playing computer games. I said that you can create your own games by using programming and in game form explain her how it can be cool. Try to make same, definitely you find what can hook {WorkerName}":result 
            }
        }
    }
    # print(Quiz["Topics"]["Carier"]["How do you think, can season influent of profit?"])
    
    quizJson = json.dumps(Quiz)
    JsonQuiz = open("Quiz.json", "w")
    JsonQuiz.write(quizJson)
    JsonQuiz.close()
    
    def Answering():
        with open("PrepareForFront.json") as fileStatic:
            print(json.load(fileStatic))




        
    Answering()

if __name__ == "__main__":
    HealthClub_Workers()
    # HealthClub_Improvements()
    HealthClub_Quiz()